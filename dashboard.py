"""Generate a visual HTML dashboard from an EZproxy SPU log file.

Usage:
    python3 dashboard.py [IP CSV file] [log file or directory] [output.html]

Example:
    python3 dashboard.py data/institutions.csv test_logs/ezproxyspu_2026_02.log dashboard.html
"""

import base64
import csv
import glob
import ipaddress
import json
import os
import re
import sys
from collections import Counter
from datetime import datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse, parse_qs

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ---------------------------------------------------------------------------
# Database name mapping
# ---------------------------------------------------------------------------

def load_database_names(json_path: str) -> dict:
    """Load the database code-to-name mapping from a JSON file.

    Returns a dict with keys 'ebsco', 'gale', and 'domains',
    each mapping codes/domains to human-readable names.
    Codes are normalized to lowercase for case-insensitive lookup.
    """
    with open(json_path, encoding='utf-8') as f:
        raw = json.load(f)

    # Normalize all codes to lowercase
    db_names = {
        'ebsco': {k.lower(): v for k, v in raw.get('ebsco', {}).items()
                  if not k.startswith('_')},
        'gale': {k.lower(): v for k, v in raw.get('gale', {}).items()
                 if not k.startswith('_')},
        'domains': {k.lower(): v for k, v in raw.get('domains', {}).items()
                    if not k.startswith('_')},
    }
    return db_names


# ---------------------------------------------------------------------------
# IP parsing (shared with ezp-analysis.py and analyze_log.py)
# ---------------------------------------------------------------------------

def parse_ip_csv(csv_path: str) -> dict:
    """Read the SharePoint-exported CSV and return {institution: [(start, end), ...]}."""
    institution_ranges = {}
    with open(csv_path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row['Institution'].strip()
            ip_text = row.get('IP Addresses', '')
            if not ip_text or not ip_text.strip():
                continue
            ranges = []
            for entry in ip_text.strip().split('\n'):
                entry = re.sub(r'\(.*?\)', '', entry).strip()
                if not re.search(r'\d+\.\d+\.\d+\.\d+', entry):
                    continue
                if ' - ' in entry:
                    parts = entry.split(' - ')
                    try:
                        start = int(ipaddress.IPv4Address(parts[0].strip()))
                        end = int(ipaddress.IPv4Address(parts[1].strip()))
                        ranges.append((start, end))
                    except (ipaddress.AddressValueError, IndexError):
                        pass
                else:
                    try:
                        ip_int = int(ipaddress.IPv4Address(entry.strip()))
                        ranges.append((ip_int, ip_int))
                    except ipaddress.AddressValueError:
                        pass
            if ranges:
                institution_ranges[name] = ranges
    return institution_ranges


def classify_ip(ip_str: str, institution_ranges: dict) -> str:
    """Return institution name or 'Off-campus' for an IP."""
    try:
        ip_int = int(ipaddress.IPv4Address(ip_str))
    except ipaddress.AddressValueError:
        return "Off-campus"
    for name, ranges in institution_ranges.items():
        for start, end in ranges:
            if start <= ip_int <= end:
                return name
    return "Off-campus"


# ---------------------------------------------------------------------------
# Log parsing
# ---------------------------------------------------------------------------

def parse_log_line(line: str) -> dict | None:
    """Parse a tab-separated SPU log line into a dictionary."""
    ts_match = re.match(r'\[(.+?)\]', line)
    if not ts_match:
        return None
    parts = [p.strip() for p in line.split('\t')]
    if len(parts) < 8:
        return None
    try:
        timestamp = datetime.strptime(ts_match.group(1), '%d/%b/%Y:%H:%M:%S %z')
    except ValueError:
        return None
    return {
        'timestamp': timestamp,
        'ip': parts[1],
        'emplid': parts[2] if parts[2] != '-' else None,  # 8-digit CUNY ID
        'session': parts[3] if parts[3] != '-' else None,
        'action': parts[4],
        'referrer': parts[5] if parts[5] != '-' else None,
        'url': parts[6],
        'status': parts[7],
    }


def extract_platform_name(url: str, db_names: dict) -> str | None:
    """Identify the vendor/platform from a URL (domain level).

    Groups all EBSCO databases under "EBSCO", all Gale under "Gale", etc.
    Returns None for infrastructure (EZproxy, Primo, DOI, Google Scholar).
    """
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower()

        # Skip infrastructure
        if 'ezproxy' in domain or 'primo' in domain or 'doi.org' in domain:
            return None
        if 'scholar.google' in domain or 'illiad' in domain:
            return None

        # Group EBSCO domains
        if any(d in domain for d in ['ebsco.com', 'ebscohost.com', 'ebsco.zone']):
            return 'EBSCO'

        # Group Gale domains
        if 'gale.com' in domain or 'galegroup.com' in domain:
            return 'Gale'

        # Clean domain
        clean = domain
        for prefix in ['www.', 'search.', 'login.', 'link.', 'go.', 'find.',
                       'openurl.', 'advance.', 'logon.']:
            if clean.startswith(prefix):
                clean = clean[len(prefix):]

        # Check domain mapping
        name = db_names['domains'].get(clean)
        if name:
            return name
        for d, n in db_names['domains'].items():
            if d in clean:
                return n

        return clean
    except Exception:
        return 'unknown'


def extract_database_name(url: str, db_names: dict) -> str:
    """Identify the specific database or resource from a URL.

    Checks for database codes in URL parameters (EBSCO db=, Gale p=),
    then falls back to domain-level identification.

    Args:
        url: The full URL from the log line
        db_names: The mapping dict from load_database_names()

    Returns:
        A human-readable database name, or the cleaned domain if no match.
    """
    try:
        parsed = urlparse(url)
        domain = parsed.netloc.lower()
        query = parse_qs(parsed.query)

        # Clean domain for matching
        clean_domain = domain
        for prefix in ['www.', 'search.', 'login.', 'link.', 'go.', 'find.',
                       'openurl.', 'advance.', 'logon.']:
            if clean_domain.startswith(prefix):
                clean_domain = clean_domain[len(prefix):]

        # --- EBSCO: look for db= or defaultdb= parameter ---
        ebsco_domains = ['ebsco.com', 'ebscohost.com', 'research.ebsco.com',
                         'search.ebscohost.com', 'openurl.ebsco.com']
        if any(d in domain for d in ebsco_domains):
            # Check db= parameter
            db_code = None
            if 'db' in query:
                db_code = query['db'][0].lower()
            elif 'defaultdb' in query:
                db_code = query['defaultdb'][0].lower()

            if db_code:
                # Handle comma-separated multi-database searches
                # (take the first one as the "primary" database)
                if ',' in db_code:
                    db_code = db_code.split(',')[0]
                name = db_names['ebsco'].get(db_code)
                if name:
                    return name
                return f"EBSCO ({db_code})"

            # EBSCO URL without a db code (login pages, OAuth, etc.)
            return None  # skip, not a specific database

        # --- EBSCO authentication domain (not a database) ---
        if 'ebsco.zone' in domain:
            return None  # skip, authentication infrastructure

        # --- Gale: look for p= parameter or /apps/pub/ path ---
        if 'gale.com' in domain or 'galegroup.com' in domain:
            product_code = None
            if 'p' in query:
                product_code = query['p'][0]
            elif 'db' in query:
                product_code = query['db'][0]
            elif '/apps/pub/' in parsed.path:
                # URL like link.gale.com/apps/pub/0ESZ/AONE?u=...
                path_parts = parsed.path.split('/')
                try:
                    pub_idx = path_parts.index('pub')
                    if pub_idx + 2 < len(path_parts):
                        product_code = path_parts[pub_idx + 2]
                except (ValueError, IndexError):
                    pass

            if product_code:
                # Skip static assets (CSS, JS, images)
                if '.' in product_code and product_code.split('.')[-1] in \
                        ['js', 'css', 'woff2', 'svg', 'png', 'jpg', 'gif']:
                    return "Gale (general)"
                name = db_names['gale'].get(product_code.lower())
                if name:
                    return name
                return f"Gale ({product_code})"

            return None  # Gale page without specific product code

        # --- Skip infrastructure pages (not databases) ---
        if 'ezproxy' in domain:
            return None  # EZproxy login pages
        if 'primo' in domain:
            return None  # discovery layer (counted separately)
        if 'doi.org' in domain:
            return None  # DOI resolver, not a database
        if 'scholar.google' in domain:
            return None  # Google Scholar, not a database
        if 'illiad' in domain:
            return None  # ILL system, not a database

        # --- Domain-level lookup ---
        domain_name = db_names['domains'].get(clean_domain)
        if domain_name:
            return domain_name

        # Partial domain match (e.g., 'kanopystreaming.com' in longer domain)
        for d, name in db_names['domains'].items():
            if d in clean_domain:
                return name

        return clean_domain

    except Exception:
        return 'unknown'


# ---------------------------------------------------------------------------
# Compute all dashboard metrics
# ---------------------------------------------------------------------------

def compute_metrics(records: list, institution_ranges: dict, log_name: str,
                    db_names: dict) -> dict:
    """Compute all metrics from parsed log records. Returns a dict for JSON embedding."""

    total = len(records)
    date_start = records[0]['timestamp'].strftime('%b %d, %Y')
    date_end = records[-1]['timestamp'].strftime('%b %d, %Y')

    # --- Single pass over records for all counters ---
    inst_counts = Counter()
    emplids = set()
    sessions = set()
    authenticated = 0
    platform_counts = Counter()
    resource_counts = Counter()
    primo_count = 0
    hour_counts = Counter()
    day_counts = Counter()
    dow_counts = Counter()
    status_counts = Counter()
    action_counts = Counter()

    for r in records:
        # Institution
        inst = classify_ip(r['ip'], institution_ranges)
        r['institution'] = inst
        inst_counts[inst] += 1

        # Users & sessions
        if r['emplid']:
            emplids.add(r['emplid'])
            authenticated += 1
        if r['session']:
            sessions.add(r['session'])

        # Platforms & databases
        platform = extract_platform_name(r['url'], db_names)
        if platform:
            platform_counts[platform] += 1
        db_name = extract_database_name(r['url'], db_names)
        if db_name:
            resource_counts[db_name] += 1

        # Primo referrals
        if r['referrer'] and 'primo' in r['referrer'].lower():
            primo_count += 1

        # Time distributions
        ts = r['timestamp']
        hour_counts[ts.hour] += 1
        day_counts[ts.strftime('%Y-%m-%d')] += 1
        dow_counts[ts.strftime('%A')] += 1

        # Status & actions
        status_counts[r['status']] += 1
        action_counts[r['action']] += 1

    # --- Aggregate results ---
    off_campus = inst_counts.pop('Off-campus', 0)
    on_campus = sum(inst_counts.values())

    inst_breakdown = [
        {'name': name, 'count': count}
        for name, count in inst_counts.most_common()
        if count > 0
    ]

    top_platforms = [
        {'name': name, 'count': count}
        for name, count in platform_counts.most_common(15)
    ]

    top_resources = [
        {'name': name, 'count': count}
        for name, count in resource_counts.most_common(20)
    ]

    hourly = [hour_counts.get(h, 0) for h in range(24)]

    sorted_days = sorted(day_counts.items())
    daily_labels = [datetime.strptime(d, '%Y-%m-%d').strftime('%b %d') for d, _ in sorted_days]
    daily_values = [c for _, c in sorted_days]

    dow_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    dow_values = [dow_counts.get(d, 0) for d in dow_order]

    status_data = [{'code': code, 'count': count}
                   for code, count in status_counts.most_common()]

    action_data = [{'action': action, 'count': count}
                   for action, count in action_counts.most_common()]

    return {
        'logName': log_name,
        'dateStart': date_start,
        'dateEnd': date_end,
        'total': total,
        'onCampus': on_campus,
        'offCampus': off_campus,
        'uniqueUsers': len(emplids),
        'uniqueSessions': len(sessions),
        'authenticated': authenticated,
        'unauthenticated': total - authenticated,
        'primoReferrals': primo_count,
        'institutionBreakdown': inst_breakdown,
        'topPlatforms': top_platforms,
        'topResources': top_resources,
        'hourly': hourly,
        'dailyLabels': daily_labels,
        'dailyValues': daily_values,
        'dowLabels': dow_order,
        'dowValues': dow_values,
        'statusCodes': status_data,
        'actionTypes': action_data,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def build_excel_data_uri(metrics: dict) -> tuple:
    """Build a formatted Excel workbook from dashboard metrics.

    Returns (data_uri_string, suggested_filename). The data URI can be
    embedded directly in an <a href="..."> tag for one-click download.
    """
    wb = Workbook()

    # -- Shared styles (12pt Arial per project convention) --
    data_font = Font(name='Arial', size=12)
    bold_font = Font(name='Arial', size=12, bold=True)
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0',
                              fill_type='solid')
    section_font = Font(name='Arial', size=14, bold=True, color='2B6CB0')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def write_table(ws, headers, rows, start_row=1):
        """Write a header row + data rows. Returns the next empty row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for r_idx, row_data in enumerate(rows, start_row + 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
        return start_row + 1 + len(rows)

    def write_section_title(ws, row, title):
        """Write a section title in a larger bold font."""
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        return row + 1

    def autofit_columns(ws):
        """Set column widths based on content length."""
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    total = metrics['total']

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = 'Summary'
    summary_rows = [
        ('Log File', metrics['logName']),
        ('Date Range', f"{metrics['dateStart']} – {metrics['dateEnd']}"),
        ('Total Connections', f"{total:,}"),
        ('Unique Users', f"{metrics['uniqueUsers']:,}"),
        ('Unique Sessions', f"{metrics['uniqueSessions']:,}"),
        ('Authenticated', f"{metrics['authenticated']:,}"),
        ('Unauthenticated', f"{metrics['unauthenticated']:,}"),
        ('On-Campus', f"{metrics['onCampus']:,}"),
        ('Off-Campus', f"{metrics['offCampus']:,}"),
        ('Primo Referrals', f"{metrics['primoReferrals']:,}"),
    ]
    for r_idx, (label, value) in enumerate(summary_rows, 1):
        lbl_cell = ws.cell(row=r_idx, column=1, value=label)
        lbl_cell.font = bold_font
        val_cell = ws.cell(row=r_idx, column=2, value=value)
        val_cell.font = data_font
    autofit_columns(ws)

    # ---- Sheet 2: Institutions ----
    ws_inst = wb.create_sheet('Institutions')
    rows = [(d['name'], d['count']) for d in metrics['institutionBreakdown']]
    write_table(ws_inst, ['Institution', 'Connections'], rows)
    autofit_columns(ws_inst)

    # ---- Sheet 3: Platforms ----
    ws_plat = wb.create_sheet('Platforms')
    rows = [(d['name'], d['count']) for d in metrics['topPlatforms']]
    write_table(ws_plat, ['Platform', 'Connections'], rows)
    autofit_columns(ws_plat)

    # ---- Sheet 4: Databases ----
    ws_db = wb.create_sheet('Databases')
    rows = [(d['name'], d['count']) for d in metrics['topResources']]
    write_table(ws_db, ['Database', 'Connections'], rows)
    autofit_columns(ws_db)

    # ---- Sheet 5: Time Patterns ----
    ws_time = wb.create_sheet('Time Patterns')
    row = 1

    # Hourly
    row = write_section_title(ws_time, row, 'Hourly Distribution')
    hour_labels = [f"{h % 12 or 12} {'am' if h < 12 else 'pm'}"
                   for h in range(24)]
    hour_rows = list(zip(hour_labels, metrics['hourly']))
    row = write_table(ws_time, ['Hour', 'Connections'], hour_rows, start_row=row)
    row += 1  # blank row separator

    # Daily
    row = write_section_title(ws_time, row, 'Daily Distribution')
    daily_rows = list(zip(metrics['dailyLabels'], metrics['dailyValues']))
    row = write_table(ws_time, ['Date', 'Connections'], daily_rows, start_row=row)
    row += 1

    # Day of week
    row = write_section_title(ws_time, row, 'Day of Week')
    dow_rows = list(zip(metrics['dowLabels'], metrics['dowValues']))
    write_table(ws_time, ['Day', 'Connections'], dow_rows, start_row=row)
    autofit_columns(ws_time)

    # ---- Sheet 6: Status & Actions ----
    ws_status = wb.create_sheet('Status & Actions')
    row = 1

    # Status codes
    row = write_section_title(ws_status, row, 'HTTP Status Codes')
    status_rows = []
    for d in metrics['statusCodes']:
        pct = f"{d['count'] / total * 100:.1f}%" if total else '0%'
        status_rows.append((d['code'], d['count'], pct))
    row = write_table(ws_status, ['Status Code', 'Count', '%'], status_rows,
                      start_row=row)
    row += 1

    # Action types
    row = write_section_title(ws_status, row, 'Action Types')
    action_rows = []
    for d in metrics['actionTypes']:
        pct = f"{d['count'] / total * 100:.1f}%" if total else '0%'
        action_rows.append((d['action'], d['count'], pct))
    write_table(ws_status, ['Action', 'Count', '%'], action_rows, start_row=row)
    autofit_columns(ws_status)

    # ---- Serialize to base64 ----
    buf = BytesIO()
    wb.save(buf)
    b64 = base64.b64encode(buf.getvalue()).decode('ascii')

    safe_name = re.sub(r'[^\w\-]', '_', metrics['logName'].replace('.log', ''))
    filename = f"ezproxy_report_{safe_name}.xlsx"
    data_uri = (
        'data:application/vnd.openxmlformats-officedocument'
        f'.spreadsheetml.sheet;base64,{b64}'
    )
    return data_uri, filename


# ---------------------------------------------------------------------------
# HTML template
# ---------------------------------------------------------------------------

_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>EZproxy Dashboard — __LOG_NAME__</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    font-family: Arial, sans-serif;
    font-size: 12pt;
    background: #f0f4f8;
    color: #1a202c;
    padding: 24px;
  }
  h1 {
    font-size: 22pt;
    color: #2b6cb0;
    margin-bottom: 4px;
  }
  .subtitle {
    color: #718096;
    font-size: 11pt;
    margin-bottom: 24px;
  }
  .cards {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
    margin-bottom: 28px;
  }
  .card {
    background: #fff;
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    text-align: center;
  }
  .card .number {
    font-size: 28pt;
    font-weight: bold;
    line-height: 1.1;
  }
  .card .label {
    font-size: 10pt;
    color: #4a5568;
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
  }
  .card.blue .number   { color: #2b6cb0; }
  .card.teal .number   { color: #2c7a7b; }
  .card.purple .number { color: #6b46c1; }
  .card.orange .number { color: #c05621; }
  .card.green .number  { color: #276749; }

  .grid-2 {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 20px;
    margin-bottom: 20px;
  }
  .grid-3 {
    display: grid;
    grid-template-columns: 1fr 1fr 1fr;
    gap: 20px;
    margin-bottom: 20px;
  }
  @media (max-width: 900px) {
    .grid-2, .grid-3 { grid-template-columns: 1fr; }
  }

  .panel {
    background: #fff;
    border-radius: 10px;
    padding: 20px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.1);
  }
  .panel h2 {
    font-size: 13pt;
    color: #2d3748;
    margin-bottom: 12px;
    border-bottom: 2px solid #e2e8f0;
    padding-bottom: 8px;
  }
  .chart-container {
    position: relative;
    width: 100%;
  }
  .chart-container.small {
    max-width: 280px;
    margin: 0 auto;
  }

  table {
    width: 100%;
    border-collapse: collapse;
    font-size: 11pt;
  }
  th {
    text-align: left;
    padding: 8px 12px;
    background: #edf2f7;
    color: #4a5568;
    font-weight: 600;
  }
  td {
    padding: 8px 12px;
    border-bottom: 1px solid #e2e8f0;
  }
  tr:hover td { background: #f7fafc; }

  .footer {
    text-align: center;
    color: #a0aec0;
    font-size: 9pt;
    margin-top: 32px;
    padding-top: 16px;
    border-top: 1px solid #e2e8f0;
  }

  .export-buttons { margin-bottom: 20px; }
  .export-buttons a {
    display: inline-block;
    font-family: Arial, sans-serif;
    font-size: 11pt;
    padding: 10px 24px;
    background: #2b6cb0;
    color: white;
    text-decoration: none;
    border-radius: 6px;
    cursor: pointer;
    margin-right: 10px;
    transition: background 0.2s;
  }
  .export-buttons a:hover { background: #2c5282; }

  @media print {
    .export-buttons { display: none; }
    body { background: white; padding: 0; }
    .panel { box-shadow: none; break-inside: avoid; }
    .card { box-shadow: none; }
    .footer { border-top: none; }
  }
</style>
</head>
<body>
<main>

<h1>EZproxy Usage Dashboard</h1>
<p class="subtitle">__LOG_NAME__ &nbsp;·&nbsp; __DATE_START__ – __DATE_END__</p>

<div class="export-buttons">
  <a href="#" role="button" onclick="window.print(); return false;">Save as PDF</a>
  <a href="__EXCEL_DATA_URI__" download="__EXCEL_FILENAME__">Export to Excel</a>
</div>

<!-- Summary Cards -->
<div class="cards">
  <div class="card blue">
    <div class="number">__TOTAL__</div>
    <div class="label">Total Connections</div>
  </div>
  <div class="card teal">
    <div class="number">__UNIQUE_USERS__</div>
    <div class="label">Unique EMPLIDs</div>
  </div>
  <div class="card purple">
    <div class="number">__UNIQUE_SESSIONS__</div>
    <div class="label">Unique Sessions</div>
  </div>
  <div class="card orange">
    <div class="number">__PRIMO_PCT__%</div>
    <div class="label">From Primo</div>
  </div>
  <div class="card green">
    <div class="number">__ON_CAMPUS_PCT__%</div>
    <div class="label">On-Campus</div>
  </div>
</div>

<!-- Row 1: Campus split + Institution breakdown -->
<div class="grid-2">
  <div class="panel">
    <h2>On-Campus vs Off-Campus</h2>
    <div class="chart-container small">
      <canvas id="campusPie" aria-label="Doughnut chart showing on-campus versus off-campus connections" role="img"></canvas>
    </div>
  </div>
  <div class="panel">
    <h2>On-Campus by Institution</h2>
    <div class="chart-container">
      <canvas id="instBar" aria-label="Bar chart showing connections by institution" role="img"></canvas>
    </div>
  </div>
</div>

<!-- Row 2: Platforms + Databases -->
<div class="grid-2">
  <div class="panel">
    <h2>Top Platforms</h2>
    <div class="chart-container">
      <canvas id="platformBar" aria-label="Bar chart showing top platforms by connection count" role="img"></canvas>
    </div>
  </div>
  <div class="panel">
    <h2>Top Databases</h2>
    <div class="chart-container">
      <canvas id="resourceBar" aria-label="Bar chart showing top databases by connection count" role="img"></canvas>
    </div>
  </div>
</div>

<!-- Row 3: Hourly + Daily -->
<div class="grid-2">
  <div class="panel">
    <h2>Connections by Hour</h2>
    <div class="chart-container">
      <canvas id="hourlyBar" aria-label="Bar chart showing connections by hour of day" role="img"></canvas>
    </div>
  </div>
  <div class="panel">
    <h2>Connections by Day</h2>
    <div class="chart-container">
      <canvas id="dailyBar" aria-label="Bar chart showing connections by date" role="img"></canvas>
    </div>
  </div>
</div>

<!-- Row 4: Day of week + Auth + Status -->
<div class="grid-3">
  <div class="panel">
    <h2>Day of Week</h2>
    <div class="chart-container">
      <canvas id="dowBar" aria-label="Bar chart showing connections by day of week" role="img"></canvas>
    </div>
  </div>
  <div class="panel">
    <h2>Authenticated vs Not</h2>
    <div class="chart-container small">
      <canvas id="authDonut" aria-label="Doughnut chart showing authenticated versus unauthenticated connections" role="img"></canvas>
    </div>
  </div>
  <div class="panel">
    <h2>HTTP Status &amp; Actions</h2>
    <table>
      <thead><tr><th scope="col">Status</th><th scope="col">Count</th><th scope="col">%</th></tr></thead>
      <tbody id="statusTable"></tbody>
    </table>
    <br>
    <table>
      <thead><tr><th scope="col">Action</th><th scope="col">Count</th><th scope="col">%</th></tr></thead>
      <tbody id="actionTable"></tbody>
    </table>
  </div>
</div>

<footer class="footer">
  Generated __GEN_DATE__ &nbsp;·&nbsp; EZproxy Analysis Dashboard
</footer>
</main>

<script>
// --- Embedded data ---
var DATA = __JSON_DATA__;

// --- Color palette ---
var BLUE   = ['#2b6cb0','#3182ce','#4299e1','#63b3ed','#90cdf4','#bee3f8'];
var TEAL   = ['#2c7a7b','#38a169','#48bb78','#68d391','#9ae6b4','#c6f6d5'];
var WARM   = ['#c05621','#dd6b20','#ed8936','#f6ad55','#fbd38d','#fefcbf'];
var PURPLE = ['#6b46c1','#805ad5','#9f7aea','#b794f4','#d6bcfa','#e9d8fd'];

// Generate enough colors for institution bars
function generateColors(n) {
  var all = [];
  var palettes = [BLUE, TEAL, PURPLE, WARM];
  for (var i = 0; i < n; i++) {
    all.push(palettes[i % palettes.length][Math.floor(i / palettes.length) % 6]);
  }
  return all;
}

// --- Campus Pie ---
new Chart(document.getElementById('campusPie'), {
  type: 'doughnut',
  data: {
    labels: ['On-Campus', 'Off-Campus'],
    datasets: [{
      data: [DATA.onCampus, DATA.offCampus],
      backgroundColor: [BLUE[0], '#e2e8f0'],
      borderWidth: 0
    }]
  },
  options: {
    plugins: {
      legend: { position: 'bottom', labels: { font: { family: 'Arial', size: 12 } } }
    }
  }
});

// --- Institution Bar ---
(function() {
  var labels = DATA.institutionBreakdown.map(function(d) { return d.name; });
  var values = DATA.institutionBreakdown.map(function(d) { return d.count; });
  var colors = generateColors(labels.length);
  new Chart(document.getElementById('instBar'), {
    type: 'bar',
    data: {
      labels: labels,
      datasets: [{ data: values, backgroundColor: colors, borderWidth: 0 }]
    },
    options: {
      indexAxis: 'y',
      plugins: { legend: { display: false } },
      scales: {
        x: { ticks: { font: { family: 'Arial', size: 11 } } },
        y: { ticks: { font: { family: 'Arial', size: 11 } } }
      }
    }
  });
})();

// --- Top Platforms ---
(function() {
  var labels = DATA.topPlatforms.map(function(d) { return d.name; });
  var values = DATA.topPlatforms.map(function(d) { return d.count; });
  new Chart(document.getElementById('platformBar'), {
    type: 'bar',
    data: {
      labels: labels,
      datasets: [{ data: values, backgroundColor: BLUE[2], borderWidth: 0 }]
    },
    options: {
      indexAxis: 'y',
      plugins: { legend: { display: false } },
      scales: {
        x: { ticks: { font: { family: 'Arial', size: 11 } } },
        y: { ticks: { font: { family: 'Arial', size: 11 } } }
      }
    }
  });
})();

// --- Top Databases ---
(function() {
  var labels = DATA.topResources.map(function(d) { return d.name; });
  var values = DATA.topResources.map(function(d) { return d.count; });
  new Chart(document.getElementById('resourceBar'), {
    type: 'bar',
    data: {
      labels: labels,
      datasets: [{ data: values, backgroundColor: TEAL[1], borderWidth: 0 }]
    },
    options: {
      indexAxis: 'y',
      plugins: { legend: { display: false } },
      scales: {
        x: { ticks: { font: { family: 'Arial', size: 11 } } },
        y: { ticks: { font: { family: 'Arial', size: 10 } } }
      }
    }
  });
})();

// --- Hourly Bar ---
(function() {
  var labels = [];
  for (var h = 0; h < 24; h++) {
    var hr = h % 12 || 12;
    labels.push(hr + (h < 12 ? 'am' : 'pm'));
  }
  new Chart(document.getElementById('hourlyBar'), {
    type: 'bar',
    data: {
      labels: labels,
      datasets: [{
        data: DATA.hourly,
        backgroundColor: BLUE[2],
        borderWidth: 0
      }]
    },
    options: {
      plugins: { legend: { display: false } },
      scales: {
        x: { ticks: { font: { family: 'Arial', size: 10 } } },
        y: { beginAtZero: true, ticks: { font: { family: 'Arial', size: 11 } } }
      }
    }
  });
})();

// --- Daily Bar ---
new Chart(document.getElementById('dailyBar'), {
  type: 'bar',
  data: {
    labels: DATA.dailyLabels,
    datasets: [{
      data: DATA.dailyValues,
      backgroundColor: PURPLE[2],
      borderWidth: 0
    }]
  },
  options: {
    plugins: { legend: { display: false } },
    scales: {
      x: { ticks: { font: { family: 'Arial', size: 10 }, maxRotation: 45 } },
      y: { beginAtZero: true, ticks: { font: { family: 'Arial', size: 11 } } }
    }
  }
});

// --- Day of Week Bar ---
new Chart(document.getElementById('dowBar'), {
  type: 'bar',
  data: {
    labels: DATA.dowLabels.map(function(d) { return d.substring(0, 3); }),
    datasets: [{
      data: DATA.dowValues,
      backgroundColor: WARM[1],
      borderWidth: 0
    }]
  },
  options: {
    plugins: { legend: { display: false } },
    scales: {
      x: { ticks: { font: { family: 'Arial', size: 11 } } },
      y: { beginAtZero: true, ticks: { font: { family: 'Arial', size: 11 } } }
    }
  }
});

// --- Auth Donut ---
new Chart(document.getElementById('authDonut'), {
  type: 'doughnut',
  data: {
    labels: ['Authenticated', 'Unauthenticated'],
    datasets: [{
      data: [DATA.authenticated, DATA.unauthenticated],
      backgroundColor: [TEAL[0], '#e2e8f0'],
      borderWidth: 0
    }]
  },
  options: {
    plugins: {
      legend: { position: 'bottom', labels: { font: { family: 'Arial', size: 12 } } }
    }
  }
});

// --- Status & Action Tables ---
(function() {
  var total = DATA.total;
  var html = '';
  DATA.statusCodes.forEach(function(d) {
    var pct = (d.count / total * 100).toFixed(1);
    html += '<tr><td>' + d.code + '</td><td>' + d.count.toLocaleString() +
            '</td><td>' + pct + '%</td></tr>';
  });
  document.getElementById('statusTable').innerHTML = html;

  html = '';
  DATA.actionTypes.forEach(function(d) {
    var pct = (d.count / total * 100).toFixed(1);
    html += '<tr><td>' + d.action + '</td><td>' + d.count.toLocaleString() +
            '</td><td>' + pct + '%</td></tr>';
  });
  document.getElementById('actionTable').innerHTML = html;
})();
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) != 4:
        print("Usage: python3 dashboard.py [IP CSV] [log file or directory] [output.html]")
        print("\nNote: data/database_names.json must exist in the same directory as this script.")
        sys.exit(1)

    csv_path = sys.argv[1]
    log_input = sys.argv[2]
    output_path = sys.argv[3]

    # Locate the database names mapping (next to this script)
    script_dir = Path(__file__).resolve().parent
    db_names_path = script_dir / 'data' / 'database_names.json'
    if not db_names_path.exists():
        print(f"Error: {db_names_path} not found.")
        print("This file maps database codes to human-readable names.")
        sys.exit(1)

    # Load database name mapping
    print(f"Loading database name mapping from {db_names_path}...")
    db_names = load_database_names(str(db_names_path))
    ebsco_count = len(db_names['ebsco'])
    gale_count = len(db_names['gale'])
    domain_count = len(db_names['domains'])
    print(f"  Loaded {ebsco_count} EBSCO + {gale_count} Gale + {domain_count} domain mappings.\n")

    # Load IP ranges
    print(f"Loading IP ranges from {csv_path}...")
    institution_ranges = parse_ip_csv(csv_path)
    print(f"  Loaded {len(institution_ranges)} institutions.\n")

    # Determine if input is a file or directory
    if os.path.isdir(log_input):
        log_files = sorted(glob.glob(os.path.join(log_input, '*.log')))
    else:
        log_files = [log_input]

    if not log_files:
        print(f"No .log files found in {log_input}")
        sys.exit(1)

    # Parse all log lines from all files
    all_records = []
    for filepath in log_files:
        print(f"  Parsing {filepath}...")
        with open(filepath, encoding='utf-8', errors='replace') as f:
            for line in f:
                parsed = parse_log_line(line.strip())
                if parsed:
                    all_records.append(parsed)

    if not all_records:
        print("No records parsed from log file(s).")
        sys.exit(1)

    # Sort by timestamp
    all_records.sort(key=lambda r: r['timestamp'])

    # Build a friendly name from the log file(s)
    if len(log_files) == 1:
        log_name = os.path.basename(log_files[0])
    else:
        log_name = f"{len(log_files)} log files from {os.path.basename(log_input)}"

    print(f"\n  Total records: {len(all_records):,}")
    print(f"  Computing metrics...")

    metrics = compute_metrics(all_records, institution_ranges, log_name, db_names)

    # Build Excel export
    print("  Building Excel export...")
    excel_data_uri, excel_filename = build_excel_data_uri(metrics)

    # Build HTML
    primo_pct = round(metrics['primoReferrals'] / metrics['total'] * 100, 1) if metrics['total'] else 0
    on_campus_pct = round(metrics['onCampus'] / metrics['total'] * 100, 1) if metrics['total'] else 0
    gen_date = datetime.now().strftime('%B %d, %Y at %I:%M %p')

    html = _HTML_TEMPLATE
    html = html.replace('__EXCEL_DATA_URI__', excel_data_uri)
    html = html.replace('__EXCEL_FILENAME__', excel_filename)
    html = html.replace('__LOG_NAME__', metrics['logName'])
    html = html.replace('__DATE_START__', metrics['dateStart'])
    html = html.replace('__DATE_END__', metrics['dateEnd'])
    html = html.replace('__TOTAL__', f"{metrics['total']:,}")
    html = html.replace('__UNIQUE_USERS__', f"{metrics['uniqueUsers']:,}")
    html = html.replace('__UNIQUE_SESSIONS__', f"{metrics['uniqueSessions']:,}")
    html = html.replace('__PRIMO_PCT__', str(primo_pct))
    html = html.replace('__ON_CAMPUS_PCT__', str(on_campus_pct))
    html = html.replace('__GEN_DATE__', gen_date)
    html = html.replace('__JSON_DATA__', json.dumps(metrics, ensure_ascii=False))

    Path(output_path).write_text(html, encoding='utf-8')
    print(f"\n  Dashboard written to: {output_path}")
    print("  Open it in a browser to view the charts.")


if __name__ == '__main__':
    main()
